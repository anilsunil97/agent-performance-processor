"""
Agent Performance Data Processor - Streamlit App
Processes agent performance CSV files with styling and exports to Excel
"""

import streamlit as st
import pandas as pd
import io
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

warnings.filterwarnings('ignore')

# Page config
st.set_page_config(
    page_title="Agent Performance Processor",
    page_icon="📊",
    layout="wide"
)

def load_and_clean_data(file):
    """Load CSV and perform initial cleaning"""
    try:
        # Read file content
        content = file.read().decode('utf-8', errors='ignore')
        lines = content.split('\n')
        
        # Find the row that contains 'USER NAME' (the actual header)
        header_row = 0
        for i, line in enumerate(lines):
            if 'USER NAME' in line.upper():
                header_row = i
                break
        
        # Store ALL rows before the header as metadata
        metadata_rows = [line + '\n' for line in lines[:header_row] if line.strip()]
        
        # Create StringIO for pandas to read
        data_content = '\n'.join(lines[header_row:])
        df = pd.read_csv(io.StringIO(data_content), on_bad_lines='skip', engine='python')
        
        # Columns to remove
        columns_to_delete = [
            'CURRENT USER GROUP', 'MOST RECENT USER GROUP', 'PAUSAVG', 'WAITAVG',
            'TALKAVG', 'DISPAVG', 'DEADAVG', 'CUSTAVG', 'ANS', 'SSMS', 'REDIAL',
            'test', 'testne', 'TestIT', 'TESTNC', 'TESTCB', 'Test22', 'DUPLICATE CALLS'
        ]
        
        # Drop columns (ignore if they don't exist)
        df = df.drop(columns=[col for col in columns_to_delete if col in df.columns], errors='ignore')
        
        # Remove last row (typically totals/summary)
        if len(df) > 0:
            df = df.iloc[:-1]
        
        return df, metadata_rows
    except Exception as e:
        st.error(f"Error loading file: {str(e)}")
        return None, None

def process_time_columns(df):
    """Calculate total pause time from PAUSE, DEAD, and DISPO columns"""
    try:
        # Convert time columns to timedelta
        df['TOTAL PAUSE'] = (
            pd.to_timedelta(df['PAUSE'], errors='coerce').fillna(pd.Timedelta(0)) +
            pd.to_timedelta(df['DEAD'], errors='coerce').fillna(pd.Timedelta(0)) +
            pd.to_timedelta(df['DISPO'], errors='coerce').fillna(pd.Timedelta(0))
        )
        
        # Format as HH:MM:SS
        df['TOTAL PAUSE'] = df['TOTAL PAUSE'].apply(
            lambda x: f"{int(x.total_seconds() // 3600):02d}:"
                      f"{int((x.total_seconds() % 3600) // 60):02d}:"
                      f"{int(x.total_seconds() % 60):02d}"
            if pd.notna(x) else "00:00:00"
        )
        
        return df
    except Exception as e:
        st.warning(f"Warning processing time columns: {str(e)}")
        return df

def reorder_and_sort(df):
    """Reorder columns and sort by total inbound calls"""
    try:
        # Convert ID to integer
        df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
        
        # Sort by total inbound calls (descending)
        if 'TOTAL INBOUND CALLS' in df.columns:
            df = df.sort_values(by='TOTAL INBOUND CALLS', ascending=False)
        
        # Reset index starting from 1
        df = df.reset_index(drop=True)
        df.index = df.index + 1
        
        # Reorder columns with ID first
        desired_columns = [
            'ID', 'USER NAME', 'CALLS', 'TIME', 'PAUSE', 'WAIT', 'TALK',
            'DISPO', 'DEAD', 'TOTAL PAUSE', 'CUSTOMER',
            'TOTAL INBOUND CALLS', 'TOTAL OUTBOUND CALLS'
        ]
        
        # Only include columns that exist
        existing_cols = [col for col in desired_columns if col in df.columns]
        df = df[existing_cols]
        
        # Add Remarks column as the last column
        df['REMARKS'] = ''
        
        # Add 'HD' in REMARKS if login hour (TIME) is less than 7 hours
        if 'TIME' in df.columns:
            for idx in df.index:
                try:
                    time_val = pd.to_timedelta(df.loc[idx, 'TIME'])
                    if time_val < pd.to_timedelta('7:00:00'):
                        df.loc[idx, 'REMARKS'] = 'HD'
                except:
                    pass
        
        return df
    except Exception as e:
        st.warning(f"Warning reordering columns: {str(e)}")
        return df

def apply_styling_to_dataframe(df):
    """Apply conditional formatting for display"""
    try:
        def color_row(row):
            styles = [''] * len(row)
            
            # Get column names
            cols = row.index.tolist()
            
            # Color TOTAL INBOUND CALLS
            if 'TOTAL INBOUND CALLS' in cols:
                idx = cols.index('TOTAL INBOUND CALLS')
                try:
                    val = float(row['TOTAL INBOUND CALLS'])
                    if val >= 70:
                        styles[idx] = 'background-color: #90EE90; color: black; font-weight: bold'
                    elif val >= 60:
                        styles[idx] = 'background-color: #FFA500; color: black; font-weight: bold'
                    elif val >= 50:
                        styles[idx] = 'background-color: #FFFF00; color: black; font-weight: bold'
                    else:
                        styles[idx] = 'background-color: #FF6B6B; color: black; font-weight: bold'
                except:
                    pass
            
            # Color TIME
            if 'TIME' in cols:
                idx = cols.index('TIME')
                try:
                    td_val = pd.to_timedelta(row['TIME'])
                    threshold_red = pd.to_timedelta('8:45:00')
                    threshold_hd = pd.to_timedelta('7:00:00')
                    
                    if td_val < threshold_hd:
                        styles[idx] = 'background-color: #FFFF00; color: black; font-weight: bold'
                    elif td_val < threshold_red:
                        styles[idx] = 'background-color: #FF6B6B; color: black; font-weight: bold'
                except:
                    pass
            
            # Color PAUSE - Dark red background with black text
            if 'PAUSE' in cols:
                idx = cols.index('PAUSE')
                try:
                    td_val = pd.to_timedelta(row['PAUSE'])
                    threshold = pd.to_timedelta('2:00:00')
                    if td_val > threshold:
                        styles[idx] = 'background-color: #DC143C; color: black; font-weight: bold'
                except:
                    pass
            
            # Color TOTAL PAUSE - Dark red background with black text
            if 'TOTAL PAUSE' in cols:
                idx = cols.index('TOTAL PAUSE')
                try:
                    td_val = pd.to_timedelta(row['TOTAL PAUSE'])
                    threshold = pd.to_timedelta('2:00:00')
                    if td_val > threshold:
                        styles[idx] = 'background-color: #DC143C; color: black; font-weight: bold'
                except:
                    pass
            
            # Color REMARKS - Yellow if HD
            if 'REMARKS' in cols:
                idx = cols.index('REMARKS')
                if str(row['REMARKS']).strip().upper() == 'HD':
                    styles[idx] = 'background-color: #FFFF00; color: black; font-weight: bold'
            
            return styles
        
        styled_df = df.style.apply(color_row, axis=1).set_properties(**{'text-align': 'center'})
        return styled_df
    except Exception as e:
        st.warning(f"Warning applying styles: {str(e)}")
        return df

def save_to_excel(df, metadata_rows):
    """Save data to Excel with metadata and styling"""
    try:
        output = io.BytesIO()
        
        # Create initial Excel file
        num_metadata_rows = len(metadata_rows)
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, startrow=num_metadata_rows, sheet_name='Agent Performance')
        
        # Load and modify workbook
        output.seek(0)
        wb = load_workbook(output)
        ws = wb.active
        
        # Get the current header row
        header_row_num = num_metadata_rows + 1
        
        # Collect all data rows
        data_rows = []
        for row in ws.iter_rows(min_row=header_row_num, values_only=False):
            data_rows.append(row)
        
        # Delete all rows from header onwards
        if ws.max_row >= header_row_num:
            ws.delete_rows(header_row_num, ws.max_row - header_row_num + 1)
        
        # Add metadata rows
        metadata_style = Font(bold=True, size=11)
        metadata_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
        
        current_row = 1
        for row in metadata_rows:
            clean_row = row.strip().replace('\n', '')
            if clean_row:
                ws.cell(row=current_row, column=1, value=clean_row)
                ws.cell(row=current_row, column=1).font = metadata_style
                ws.cell(row=current_row, column=1).fill = metadata_fill
                current_row += 1
        
        # Add data table with styling
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        header_font = Font(bold=True, color='000000')
        
        # Color definitions for conditional formatting
        green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
        orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        dark_red_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')  # Dark red (Crimson)
        black_font = Font(bold=True, color='000000')
        
        header_row_idx = None
        for idx, row_data in enumerate(data_rows):
            is_header_row = (idx == 0)
            if is_header_row:
                header_row_idx = current_row
            
            for col_idx, cell in enumerate(row_data, start=1):
                new_cell = ws.cell(row=current_row, column=col_idx)
                if cell.value is not None:
                    new_cell.value = cell.value
                
                if is_header_row:
                    new_cell.fill = header_fill
                    new_cell.font = header_font
                else:
                    # Get column name from header
                    col_name = ws.cell(row=header_row_idx, column=col_idx).value
                    
                    # Apply conditional formatting
                    if col_name == 'TOTAL INBOUND CALLS':
                        try:
                            val = float(new_cell.value)
                            if val >= 70:
                                new_cell.fill = green_fill
                                new_cell.font = black_font
                            elif val >= 60:
                                new_cell.fill = orange_fill
                                new_cell.font = black_font
                            elif val >= 50:
                                new_cell.fill = yellow_fill
                                new_cell.font = black_font
                            else:
                                new_cell.fill = red_fill
                                new_cell.font = black_font
                        except:
                            pass
                    
                    elif col_name == 'TIME':
                        try:
                            td_val = pd.to_timedelta(new_cell.value)
                            threshold_red = pd.to_timedelta('8:45:00')
                            threshold_hd = pd.to_timedelta('7:00:00')
                            
                            if td_val < threshold_hd:
                                new_cell.fill = yellow_fill
                                new_cell.font = black_font
                            elif td_val < threshold_red:
                                new_cell.fill = red_fill
                                new_cell.font = black_font
                        except:
                            pass
                    
                    elif col_name in ['PAUSE', 'TOTAL PAUSE']:
                        try:
                            td_val = pd.to_timedelta(new_cell.value)
                            threshold = pd.to_timedelta('2:00:00')
                            if td_val > threshold:
                                new_cell.fill = dark_red_fill  # Dark red background
                                new_cell.font = black_font  # Black text
                        except:
                            pass
                    
                    elif col_name == 'REMARKS':
                        if str(new_cell.value).strip().upper() == 'HD':
                            new_cell.fill = yellow_fill  # Yellow background for HD
                            new_cell.font = black_font
                
                new_cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # Add summary row below the table
        summary_row = current_row + 1
        
        # Calculate totals
        total_inbound = int(df['TOTAL INBOUND CALLS'].sum())
        avg_inbound = float(df['TOTAL INBOUND CALLS'].mean())
        
        # Style for summary
        summary_font = Font(bold=True, size=12, color='FFFFFF')
        summary_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add summary labels and values
        ws.cell(row=summary_row, column=1, value='TOTAL INBOUND CALLS')
        ws.cell(row=summary_row, column=1).font = summary_font
        ws.cell(row=summary_row, column=1).fill = summary_fill
        ws.cell(row=summary_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=summary_row, column=1).border = border
        
        ws.cell(row=summary_row, column=2, value=total_inbound)
        ws.cell(row=summary_row, column=2).font = summary_font
        ws.cell(row=summary_row, column=2).fill = summary_fill
        ws.cell(row=summary_row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=summary_row, column=2).border = border
        
        ws.cell(row=summary_row + 1, column=1, value='AVERAGE INBOUND CALLS')
        ws.cell(row=summary_row + 1, column=1).font = summary_font
        ws.cell(row=summary_row + 1, column=1).fill = summary_fill
        ws.cell(row=summary_row + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=summary_row + 1, column=1).border = border
        
        ws.cell(row=summary_row + 1, column=2, value=round(avg_inbound, 2))
        ws.cell(row=summary_row + 1, column=2).font = summary_font
        ws.cell(row=summary_row + 1, column=2).fill = summary_fill
        ws.cell(row=summary_row + 1, column=2).alignment = Alignment(horizontal='center', vertical='center')
        ws.cell(row=summary_row + 1, column=2).border = border
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output
    except Exception as e:
        st.error(f"Error creating Excel file: {str(e)}")
        return None

# Main Streamlit App
def main():
    st.title("📊 Agent Performance Data Processor - Updated")
    st.markdown("---")
    
    # File uploader
    uploaded_file = st.file_uploader(
        "Upload your Agent Performance CSV file",
        type=['csv'],
        help="Select the CSV file containing agent performance data"
    )
    
    if uploaded_file is not None:
        try:
            with st.spinner('Processing data...'):
                # Load and process data
                df, metadata_rows = load_and_clean_data(uploaded_file)
                
                if df is None:
                    st.error("Failed to load data. Please check your CSV file format.")
                    return
                
                df = process_time_columns(df)
                df = reorder_and_sort(df)
                
                # Display metadata
                st.subheader("📋 File Information")
                
                # Show metadata rows above the table
                if metadata_rows:
                    st.markdown("### Original File Header Rows:")
                    metadata_df = pd.DataFrame({
                        'Row': [f"Header {i+1}" for i in range(len(metadata_rows))],
                        'Content': [row.strip() for row in metadata_rows]
                    })
                    st.table(metadata_df)
                
                # Display summary
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Total Agents", len(df))
                with col2:
                    st.metric("Total Inbound Calls", f"{int(df['TOTAL INBOUND CALLS'].sum()):,}")
                with col3:
                    st.metric("Avg Inbound Calls", f"{df['TOTAL INBOUND CALLS'].mean():.2f}")
                
                # Display top performer
                if len(df) > 0:
                    st.success(f"Top Performer: **{df.iloc[0]['USER NAME']}** with **{int(df.iloc[0]['TOTAL INBOUND CALLS'])}** calls")
                
                st.markdown("---")
                
                # Display styled dataframe
                st.subheader("📊 Processed Data with Color Formatting")
                st.markdown("""
                **Color Legend:**
                - **Green**: >=70 calls | **Orange**: 60-69 calls | **Yellow**: 50-59 calls | **Red**: <50 calls
                - **Yellow TIME**: <7 hours (HD) | **Red TIME**: <8:45 hours
                - **Dark Red PAUSE**: >2 hours (black text)
                - **Yellow REMARKS**: Shows "HD" for login hours <7
                """)
                
                # Apply styling and display
                styled_df = apply_styling_to_dataframe(df)
                
                # Use HTML rendering for better color support
                html = styled_df.to_html(escape=False)
                st.markdown(html, unsafe_allow_html=True)
                
                # Generate Excel file
                excel_file = save_to_excel(df, metadata_rows)
                
                if excel_file:
                    # Download buttons
                    st.markdown("---")
                    st.subheader("📥 Download Processed Files")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # CSV Download
                        csv_buffer = io.StringIO()
                        if metadata_rows:
                            for row in metadata_rows:
                                csv_buffer.write(row)
                        df.to_csv(csv_buffer, index=False)
                        csv_data = csv_buffer.getvalue()
                        
                        st.download_button(
                            label="Download CSV",
                            data=csv_data,
                            file_name="cleaned_agent_performance.csv",
                            mime="text/csv"
                        )
                    
                    with col2:
                        # Excel Download
                        st.download_button(
                            label="Download Styled Excel",
                            data=excel_file,
                            file_name="styled_agent_performance.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    
                    st.success("Processing complete! Download your files above.")
                
        except Exception as e:
            st.error(f"Error processing file: {str(e)}")
            with st.expander("Show error details"):
                st.exception(e)
    
    else:
        st.info("Please upload a CSV file to get started")
        
        # Instructions
        st.markdown("---")
        st.subheader("Instructions")
        st.markdown("""
        1. Upload your agent performance CSV file
        2. The app will automatically process and format the data
        3. Download the cleaned CSV or styled Excel file
        

if __name__ == "__main__":
    main()
