"""
Agent Performance Data Processor - Native Windows GUI
Offline executable version with all functionality
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import io
import os
from pathlib import Path
import threading
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class AgentPerformanceGUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Agent Performance Data Processor")
        
        # Get screen dimensions
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # Calculate window size (80% of screen size, but with reasonable limits)
        window_width = min(1000, int(screen_width * 0.8))
        window_height = min(700, int(screen_height * 0.8))
        
        # Calculate position to center the window
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # Set window size and position
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.root.minsize(600, 400)  # Smaller minimum size
        
        # Make window resizable
        self.root.resizable(True, True)
        
        # Variables
        self.df = None
        self.metadata_rows = []
        self.processed_df = None
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the user interface"""
        # Main frame (reduced padding for smaller screens)
        main_frame = ttk.Frame(self.root, padding="5")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(3, weight=1)  # Updated for logo frame
        
        # Logo Frame (more compact)
        logo_frame = ttk.Frame(main_frame, padding="5")
        logo_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        logo_frame.columnconfigure(0, weight=1)
        
        # Main Logo Title (smaller font for compact display)
        logo_title = ttk.Label(
            logo_frame,
            text="📊 Agent Performance Processor",
            font=("Arial", 16, "bold"),
            foreground="#2c3e50"
        )
        logo_title.grid(row=0, column=0, pady=(0, 2))
        
        # Subtitle (smaller font)
        subtitle = ttk.Label(
            logo_frame,
            text="Professional Data Analysis & Reporting Tool",
            font=("Arial", 9, "italic"),
            foreground="#7f8c8d"
        )
        subtitle.grid(row=1, column=0, pady=(0, 5))
        
        # Separator (more compact)
        separator = ttk.Separator(logo_frame, orient='horizontal')
        separator.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=(2, 5))
        
        # File selection frame
        # File selection frame (reduced padding)
        file_frame = ttk.LabelFrame(main_frame, text="File Selection", padding="5")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 5))
        file_frame.columnconfigure(1, weight=1)
        
        # File selection
        ttk.Label(file_frame, text="CSV File:").grid(row=0, column=0, sticky=tk.W, padx=(0, 10))
        
        self.file_var = tk.StringVar()
        file_entry = ttk.Entry(file_frame, textvariable=self.file_var, state="readonly")
        file_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 10))
        
        # Colorful Browse button
        browse_btn = tk.Button(
            file_frame, 
            text="📁 Browse", 
            command=self.browse_file,
            bg="#3498db",  # Blue background
            fg="white",    # White text
            font=("Arial", 10, "bold"),
            relief="raised",
            bd=2,
            padx=15,
            pady=8,
            cursor="hand2"
        )
        browse_btn.grid(row=0, column=2, padx=(5, 5))
        
        # Add hover effects
        def on_browse_enter(e):
            browse_btn.config(bg="#2980b9")
        def on_browse_leave(e):
            browse_btn.config(bg="#3498db")
        browse_btn.bind("<Enter>", on_browse_enter)
        browse_btn.bind("<Leave>", on_browse_leave)
        
        # Colorful Process Data button
        process_btn = tk.Button(
            file_frame, 
            text="⚡ Process Data", 
            command=self.process_data,
            bg="#27ae60",  # Green background
            fg="white",    # White text
            font=("Arial", 10, "bold"),
            relief="raised",
            bd=2,
            padx=15,
            pady=8,
            cursor="hand2"
        )
        process_btn.grid(row=0, column=3, padx=(5, 0))
        
        # Add hover effects
        def on_process_enter(e):
            process_btn.config(bg="#229954")
        def on_process_leave(e):
            process_btn.config(bg="#27ae60")
        process_btn.bind("<Enter>", on_process_enter)
        process_btn.bind("<Leave>", on_process_leave)
        
        # Notebook for tabs (reduced padding)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 5))
        
        # Data tab
        self.data_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.data_frame, text="Processed Data")
        
        # Create treeview for data display
        self.setup_data_view()
        
        # Summary tab
        self.summary_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.summary_frame, text="Summary")
        self.setup_summary_view()
        
        # Log tab
        self.log_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.log_frame, text="Log")
        self.setup_log_view()
        
        # Export frame (reduced padding)
        export_frame = ttk.LabelFrame(main_frame, text="Export Options", padding="5")
        export_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E))
        
        # Colorful Export CSV button
        export_csv_btn = tk.Button(
            export_frame, 
            text="📄 Export CSV", 
            command=self.export_csv,
            bg="#e67e22",  # Orange background
            fg="white",    # White text
            font=("Arial", 11, "bold"),
            relief="raised",
            bd=2,
            padx=20,
            pady=10,
            cursor="hand2"
        )
        export_csv_btn.grid(row=0, column=0, padx=(0, 15))
        
        # Add hover effects
        def on_csv_enter(e):
            export_csv_btn.config(bg="#d35400")
        def on_csv_leave(e):
            export_csv_btn.config(bg="#e67e22")
        export_csv_btn.bind("<Enter>", on_csv_enter)
        export_csv_btn.bind("<Leave>", on_csv_leave)
        
        # Colorful Export Excel button
        export_excel_btn = tk.Button(
            export_frame, 
            text="📊 Export Styled Excel", 
            command=self.export_excel,
            bg="#9b59b6",  # Purple background
            fg="white",    # White text
            font=("Arial", 11, "bold"),
            relief="raised",
            bd=2,
            padx=20,
            pady=10,
            cursor="hand2"
        )
        export_excel_btn.grid(row=0, column=1, padx=(0, 15))
        
        # Add hover effects
        def on_excel_enter(e):
            export_excel_btn.config(bg="#8e44ad")
        def on_excel_leave(e):
            export_excel_btn.config(bg="#9b59b6")
        export_excel_btn.bind("<Enter>", on_excel_enter)
        export_excel_btn.bind("<Leave>", on_excel_leave)
        
        # Test Dialog button (smaller, different color)
        test_dialog_btn = tk.Button(
            export_frame, 
            text="🔧 Test Dialog", 
            command=self.test_dialog,
            bg="#95a5a6",  # Gray background
            fg="white",    # White text
            font=("Arial", 9, "bold"),
            relief="raised",
            bd=2,
            padx=15,
            pady=8,
            cursor="hand2"
        )
        test_dialog_btn.grid(row=0, column=2)
        
        # Add hover effects
        def on_test_enter(e):
            test_dialog_btn.config(bg="#7f8c8d")
        def on_test_leave(e):
            test_dialog_btn.config(bg="#95a5a6")
        test_dialog_btn.bind("<Enter>", on_test_enter)
        test_dialog_btn.bind("<Leave>", on_test_leave)
        
        # Status bar (reduced padding)
        self.status_var = tk.StringVar()
        self.status_var.set("Ready - Select a CSV file to begin")
        status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        status_bar.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
    def setup_data_view(self):
        """Setup the data view tab"""
        # Frame for treeview and scrollbars (reduced padding)
        tree_frame = ttk.Frame(self.data_frame)
        tree_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Treeview
        self.tree = ttk.Treeview(tree_frame)
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        h_scrollbar = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        v_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        h_scrollbar.grid(row=1, column=0, sticky=(tk.W, tk.E))
        
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        
    def setup_summary_view(self):
        """Setup the summary view tab"""
        summary_text = scrolledtext.ScrolledText(self.summary_frame, wrap=tk.WORD)
        summary_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.summary_text = summary_text
        
    def setup_log_view(self):
        """Setup the log view tab"""
        log_text = scrolledtext.ScrolledText(self.log_frame, wrap=tk.WORD)
        log_text.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        self.log_text = log_text
        
    def log(self, message):
        """Add message to log"""
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.root.update()
        
    def browse_file(self):
        """Browse for CSV file"""
        # Ensure dialog appears on top
        self.root.lift()
        self.root.focus_force()
        
        filename = filedialog.askopenfilename(
            parent=self.root,
            title="Select Agent Performance CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=os.path.expanduser("~/Desktop")
        )
        if filename:
            self.file_var.set(filename)
            self.log(f"Selected file: {filename}")
            
    def process_data(self):
        """Process the selected CSV file"""
        if not self.file_var.get():
            messagebox.showerror("Error", "Please select a CSV file first")
            return
            
        # Run processing in thread to prevent UI freezing
        threading.Thread(target=self._process_data_thread, daemon=True).start()
        
    def _process_data_thread(self):
        """Process data in background thread"""
        try:
            self.status_var.set("Processing data...")
            self.log("Starting data processing...")
            
            # Load and clean data
            self.df, self.metadata_rows = self.load_and_clean_data(self.file_var.get())
            if self.df is None:
                return
                
            # Process time columns
            self.df = self.process_time_columns(self.df)
            
            # Reorder and sort
            self.processed_df = self.reorder_and_sort(self.df)
            
            # Update UI in main thread
            self.root.after(0, self.update_ui_after_processing)
            
        except Exception as e:
            self.root.after(0, lambda: self.show_error(f"Error processing data: {str(e)}"))
            
    def update_ui_after_processing(self):
        """Update UI after data processing is complete"""
        try:
            # Update treeview
            self.update_treeview()
            
            # Update summary
            self.update_summary()
            
            # Switch to data tab
            self.notebook.select(0)
            
            self.status_var.set("Data processed successfully")
            self.log("Data processing completed successfully")
            
        except Exception as e:
            self.show_error(f"Error updating UI: {str(e)}")
            
    def load_and_clean_data(self, file_path):
        """Load CSV and perform initial cleaning"""
        try:
            self.log("Loading CSV file...")
            
            # Read file content
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read()
            
            lines = content.split('\n')
            
            # Find the row that contains 'USER NAME' (the actual header)
            header_row = 0
            for i, line in enumerate(lines):
                if 'USER NAME' in line.upper():
                    header_row = i
                    break
            
            # Store ALL rows before the header as metadata
            metadata_rows = [line + '\n' for line in lines[:header_row] if line.strip()]
            
            # Create StringIO for pandas to read
            data_content = '\n'.join(lines[header_row:])
            df = pd.read_csv(io.StringIO(data_content), on_bad_lines='skip', engine='python')
            
            self.log(f"Loaded {len(df)} rows of data")
            
            # Columns to remove
            columns_to_delete = [
                'CURRENT USER GROUP', 'MOST RECENT USER GROUP', 'PAUSAVG', 'WAITAVG',
                'TALKAVG', 'DISPAVG', 'DEADAVG', 'CUSTAVG', 'ANS', 'SSMS', 'REDIAL',
                'test', 'testne', 'TestIT', 'TESTNC', 'TESTCB', 'Test22', 'DUPLICATE CALLS'
            ]
            
            # Drop columns (ignore if they don't exist)
            df = df.drop(columns=[col for col in columns_to_delete if col in df.columns], errors='ignore')
            
            # Remove last row (typically totals/summary)
            if len(df) > 0:
                df = df.iloc[:-1]
            
            self.log("Data cleaning completed")
            return df, metadata_rows
            
        except Exception as e:
            self.log(f"Error loading file: {str(e)}")
            self.root.after(0, lambda: messagebox.showerror("Error", f"Error loading file: {str(e)}"))
            return None, None
            
    def process_time_columns(self, df):
        """Calculate total pause time from PAUSE, DEAD, and DISPO columns"""
        try:
            self.log("Processing time columns...")
            
            # Convert time columns to timedelta
            df['TOTAL PAUSE'] = (
                pd.to_timedelta(df['PAUSE'], errors='coerce').fillna(pd.Timedelta(0)) +
                pd.to_timedelta(df['DEAD'], errors='coerce').fillna(pd.Timedelta(0)) +
                pd.to_timedelta(df['DISPO'], errors='coerce').fillna(pd.Timedelta(0))
            )
            
            # Format as HH:MM:SS
            df['TOTAL PAUSE'] = df['TOTAL PAUSE'].apply(
                lambda x: f"{int(x.total_seconds() // 3600):02d}:"
                          f"{int((x.total_seconds() % 3600) // 60):02d}:"
                          f"{int(x.total_seconds() % 60):02d}"
                if pd.notna(x) else "00:00:00"
            )
            
            return df
            
        except Exception as e:
            self.log(f"Warning processing time columns: {str(e)}")
            return df
            
    def reorder_and_sort(self, df):
        """Reorder columns and sort by total inbound calls"""
        try:
            self.log("Reordering and sorting data...")
            
            # Convert ID to integer
            df['ID'] = pd.to_numeric(df['ID'], errors='coerce').fillna(0).astype(int)
            
            # Sort by total inbound calls (descending)
            if 'TOTAL INBOUND CALLS' in df.columns:
                df = df.sort_values(by='TOTAL INBOUND CALLS', ascending=False)
            
            # Reset index starting from 1
            df = df.reset_index(drop=True)
            df.index = df.index + 1
            
            # Reorder columns with ID first
            desired_columns = [
                'ID', 'USER NAME', 'CALLS', 'TIME', 'PAUSE', 'WAIT', 'TALK',
                'DISPO', 'DEAD', 'TOTAL PAUSE', 'CUSTOMER',
                'TOTAL INBOUND CALLS', 'TOTAL OUTBOUND CALLS'
            ]
            
            # Only include columns that exist
            existing_cols = [col for col in desired_columns if col in df.columns]
            df = df[existing_cols]
            
            # Add Remarks column as the last column
            df['REMARKS'] = ''
            
            # Add 'HD' in REMARKS if login hour (TIME) is less than 7 hours
            if 'TIME' in df.columns:
                for idx in df.index:
                    try:
                        time_val = pd.to_timedelta(df.loc[idx, 'TIME'])
                        if time_val < pd.to_timedelta('7:00:00'):
                            df.loc[idx, 'REMARKS'] = 'HD'
                    except:
                        pass
            
            return df
            
        except Exception as e:
            self.log(f"Warning reordering columns: {str(e)}")
            return df
            
    def update_treeview(self):
        """Update the treeview with processed data using exact Streamlit colors"""
        if self.processed_df is None:
            return
            
        # Clear existing data
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        # Configure columns
        columns = list(self.processed_df.columns)
        self.tree['columns'] = columns
        self.tree['show'] = 'headings'
        
        # Configure column headings and widths
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=100, minwidth=50)
        
        # Define tags for different colors (matching Streamlit exactly)
        self.tree.tag_configure('excellent', background='#90EE90', foreground='black')  # Light green
        self.tree.tag_configure('good', background='#FFA500', foreground='black')       # Orange
        self.tree.tag_configure('average', background='#FFFF00', foreground='black')    # Yellow
        self.tree.tag_configure('below_avg', background='#FF6B6B', foreground='black')  # Red
        self.tree.tag_configure('time_hd', background='#FFFF00', foreground='black')    # Yellow for HD time
        self.tree.tag_configure('time_red', background='#FF6B6B', foreground='black')   # Red for <8:45 time
        self.tree.tag_configure('pause_high', background='#DC143C', foreground='black') # Dark red for >2hr pause
        self.tree.tag_configure('remarks_hd', background='#FFFF00', foreground='black') # Yellow for HD remarks
            
        # Insert data with exact Streamlit styling
        for index, row in self.processed_df.iterrows():
            values = [str(row[col]) for col in columns]
            
            # Determine row styling based on TOTAL INBOUND CALLS
            tag = ''
            if 'TOTAL INBOUND CALLS' in columns:
                try:
                    calls = float(row['TOTAL INBOUND CALLS'])
                    if calls >= 70:
                        tag = 'excellent'
                    elif calls >= 60:
                        tag = 'good'
                    elif calls >= 50:
                        tag = 'average'
                    else:
                        tag = 'below_avg'
                except:
                    pass
            
            # Check for TIME coloring
            if 'TIME' in columns:
                try:
                    td_val = pd.to_timedelta(row['TIME'])
                    threshold_red = pd.to_timedelta('8:45:00')
                    threshold_hd = pd.to_timedelta('7:00:00')
                    
                    if td_val < threshold_hd:
                        tag = 'time_hd'
                    elif td_val < threshold_red:
                        tag = 'time_red'
                except:
                    pass
            
            # Check for PAUSE/TOTAL PAUSE coloring
            for pause_col in ['PAUSE', 'TOTAL PAUSE']:
                if pause_col in columns:
                    try:
                        td_val = pd.to_timedelta(row[pause_col])
                        threshold = pd.to_timedelta('2:00:00')
                        if td_val > threshold:
                            tag = 'pause_high'
                    except:
                        pass
            
            # Check for REMARKS HD
            if 'REMARKS' in columns and str(row['REMARKS']).strip().upper() == 'HD':
                tag = 'remarks_hd'
            
            item = self.tree.insert('', 'end', values=values, tags=(tag,))
                    
    def update_summary(self):
        """Update the summary tab"""
        if self.processed_df is None:
            return
            
        summary = []
        summary.append("📊 AGENT PERFORMANCE SUMMARY")
        summary.append("=" * 50)
        summary.append("")
        
        # Basic statistics
        total_agents = len(self.processed_df)
        total_inbound = int(self.processed_df['TOTAL INBOUND CALLS'].sum())
        avg_inbound = self.processed_df['TOTAL INBOUND CALLS'].mean()
        
        summary.append(f"📈 Total Agents: {total_agents}")
        summary.append(f"📞 Total Inbound Calls: {total_inbound:,}")
        summary.append(f"📊 Average Inbound Calls: {avg_inbound:.2f}")
        summary.append("")
        
        # Top performer
        if len(self.processed_df) > 0:
            top_performer = self.processed_df.iloc[0]
            summary.append(f"🏆 Top Performer: {top_performer['USER NAME']}")
            summary.append(f"   Calls: {int(top_performer['TOTAL INBOUND CALLS'])}")
            summary.append("")
        
        # Performance distribution
        summary.append("📊 PERFORMANCE DISTRIBUTION:")
        summary.append("-" * 30)
        
        excellent = len(self.processed_df[self.processed_df['TOTAL INBOUND CALLS'] >= 70])
        good = len(self.processed_df[(self.processed_df['TOTAL INBOUND CALLS'] >= 60) & 
                                   (self.processed_df['TOTAL INBOUND CALLS'] < 70)])
        average = len(self.processed_df[(self.processed_df['TOTAL INBOUND CALLS'] >= 50) & 
                                      (self.processed_df['TOTAL INBOUND CALLS'] < 60)])
        below_avg = len(self.processed_df[self.processed_df['TOTAL INBOUND CALLS'] < 50])
        
        summary.append(f"🟢 Excellent (≥70 calls): {excellent} agents")
        summary.append(f"🟠 Good (60-69 calls): {good} agents")
        summary.append(f"🟡 Average (50-59 calls): {average} agents")
        summary.append(f"🔴 Below Average (<50 calls): {below_avg} agents")
        summary.append("")
        
        # HD (Half Day) analysis
        hd_count = len(self.processed_df[self.processed_df['REMARKS'] == 'HD'])
        summary.append(f"🟡 Half Day (HD) Agents: {hd_count}")
        summary.append("")
        
        # Color legend
        summary.append("🎨 COLOR LEGEND:")
        summary.append("-" * 20)
        summary.append("🟢 Green: ≥70 calls (Excellent)")
        summary.append("🟠 Orange: 60-69 calls (Good)")
        summary.append("🟡 Yellow: 50-59 calls (Average)")
        summary.append("🔴 Red: <50 calls (Below Average)")
        summary.append("🟡 HD: Login time <7 hours")
        
        # Update summary text
        self.summary_text.delete(1.0, tk.END)
        self.summary_text.insert(1.0, '\n'.join(summary))
        
    def test_dialog(self):
        """Test if file dialog works"""
        try:
            self.log("Testing file dialog...")
            self.root.lift()
            self.root.focus_force()
            
            filename = filedialog.asksaveasfilename(
                parent=self.root,
                title="Test Save Dialog",
                defaultextension=".txt",
                filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
            )
            
            if filename:
                self.log(f"Dialog test successful! Selected: {filename}")
                messagebox.showinfo("Test Result", f"Dialog works! Selected: {filename}")
            else:
                self.log("Dialog test: User cancelled")
                messagebox.showinfo("Test Result", "Dialog appeared but user cancelled")
                
        except Exception as e:
            error_msg = f"Dialog test failed: {str(e)}"
            self.log(f"ERROR: {error_msg}")
            messagebox.showerror("Test Failed", error_msg)
            
    def export_csv(self):
        """Export data to CSV"""
        if self.processed_df is None:
            messagebox.showerror("Error", "No data to export. Please process a file first.")
            return
            
        try:
            # Ensure dialog appears on top
            self.root.lift()
            self.root.focus_force()
            self.root.update()
            
            # Try different approaches for the dialog
            try:
                filename = filedialog.asksaveasfilename(
                    parent=self.root,
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                    title="Save CSV File",
                    initialdir=os.path.expanduser("~/Desktop")
                )
            except:
                # Fallback without initialdir
                filename = filedialog.asksaveasfilename(
                    parent=self.root,
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                    title="Save CSV File"
                )
            
            self.log(f"Dialog returned filename: {filename}")
            
            if filename:
                # Create CSV with metadata
                with open(filename, 'w', newline='', encoding='utf-8') as f:
                    # Write metadata
                    for row in self.metadata_rows:
                        f.write(row)
                    
                    # Write data
                    self.processed_df.to_csv(f, index=False)
                
                messagebox.showinfo("Success", f"Data exported to {filename}")
                self.log(f"Data exported to CSV: {filename}")
            else:
                self.log("Export cancelled by user")
                
        except Exception as e:
            error_msg = f"Error exporting CSV: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log(f"ERROR: {error_msg}")
                
    def export_excel(self):
        """Export data to styled Excel"""
        if self.processed_df is None:
            messagebox.showerror("Error", "No data to export. Please process a file first.")
            return
            
        try:
            # Ensure dialog appears on top
            self.root.lift()
            self.root.focus_force()
            self.root.update()
            
            # Try different approaches for the dialog
            try:
                filename = filedialog.asksaveasfilename(
                    parent=self.root,
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Save Excel File",
                    initialdir=os.path.expanduser("~/Desktop")
                )
            except:
                # Fallback without initialdir
                filename = filedialog.asksaveasfilename(
                    parent=self.root,
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    title="Save Excel File"
                )
            
            self.log(f"Dialog returned filename: {filename}")
            
            if filename:
                self.status_var.set("Creating Excel file...")
                self.log("Creating styled Excel file...")
                
                # Run export in thread
                threading.Thread(
                    target=self._export_excel_thread, 
                    args=(filename,), 
                    daemon=True
                ).start()
            else:
                self.log("Export cancelled by user")
                
        except Exception as e:
            error_msg = f"Error exporting Excel: {str(e)}"
            messagebox.showerror("Error", error_msg)
            self.log(f"ERROR: {error_msg}")
                
    def _export_excel_thread(self, filename):
        """Export Excel in background thread with exact Streamlit app styling"""
        try:
            output = io.BytesIO()
            
            # Create initial Excel file
            num_metadata_rows = len(self.metadata_rows)
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                self.processed_df.to_excel(writer, index=False, startrow=num_metadata_rows, sheet_name='Agent Performance')
            
            # Load and modify workbook
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active
            
            # Get the current header row
            header_row_num = num_metadata_rows + 1
            
            # Collect all data rows
            data_rows = []
            for row in ws.iter_rows(min_row=header_row_num, values_only=False):
                data_rows.append(row)
            
            # Delete all rows from header onwards
            if ws.max_row >= header_row_num:
                ws.delete_rows(header_row_num, ws.max_row - header_row_num + 1)
            
            # Add metadata rows
            metadata_style = Font(bold=True, size=11)
            metadata_fill = PatternFill(start_color='E8F4F8', end_color='E8F4F8', fill_type='solid')
            
            current_row = 1
            for row in self.metadata_rows:
                clean_row = row.strip().replace('\n', '')
                if clean_row:
                    ws.cell(row=current_row, column=1, value=clean_row)
                    ws.cell(row=current_row, column=1).font = metadata_style
                    ws.cell(row=current_row, column=1).fill = metadata_fill
                    current_row += 1
            
            # Add data table with styling
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            
            # Color definitions for conditional formatting (EXACT SAME AS STREAMLIT)
            green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
            dark_red_fill = PatternFill(start_color='DC143C', end_color='DC143C', fill_type='solid')
            black_font = Font(bold=True, color='000000')
            
            header_row_idx = None
            for idx, row_data in enumerate(data_rows):
                is_header_row = (idx == 0)
                if is_header_row:
                    header_row_idx = current_row
                
                for col_idx, cell in enumerate(row_data, start=1):
                    new_cell = ws.cell(row=current_row, column=col_idx)
                    if cell.value is not None:
                        new_cell.value = cell.value
                    
                    if is_header_row:
                        new_cell.fill = header_fill
                        new_cell.font = header_font
                    else:
                        # Get column name from header
                        col_name = ws.cell(row=header_row_idx, column=col_idx).value
                        
                        # Apply conditional formatting (EXACT SAME AS STREAMLIT)
                        if col_name == 'TOTAL INBOUND CALLS':
                            try:
                                val = float(new_cell.value)
                                if val >= 70:
                                    new_cell.fill = green_fill
                                    new_cell.font = black_font
                                elif val >= 60:
                                    new_cell.fill = orange_fill
                                    new_cell.font = black_font
                                elif val >= 50:
                                    new_cell.fill = yellow_fill
                                    new_cell.font = black_font
                                else:
                                    new_cell.fill = red_fill
                                    new_cell.font = black_font
                            except:
                                pass
                        
                        elif col_name == 'TIME':
                            try:
                                td_val = pd.to_timedelta(new_cell.value)
                                threshold_red = pd.to_timedelta('8:45:00')
                                threshold_hd = pd.to_timedelta('7:00:00')
                                
                                if td_val < threshold_hd:
                                    new_cell.fill = yellow_fill
                                    new_cell.font = black_font
                                elif td_val < threshold_red:
                                    new_cell.fill = red_fill
                                    new_cell.font = black_font
                            except:
                                pass
                        
                        elif col_name in ['PAUSE', 'TOTAL PAUSE']:
                            try:
                                td_val = pd.to_timedelta(new_cell.value)
                                threshold = pd.to_timedelta('2:00:00')
                                if td_val > threshold:
                                    new_cell.fill = dark_red_fill
                                    new_cell.font = black_font
                            except:
                                pass
                        
                        elif col_name == 'REMARKS':
                            if str(new_cell.value).strip().upper() == 'HD':
                                new_cell.fill = yellow_fill
                                new_cell.font = black_font
                    
                    new_cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
            
            # Add summary rows below the table (separate rows)
            summary_row_1 = current_row + 2  # Leave one empty row
            summary_row_2 = current_row + 3
            
            # Calculate totals
            total_inbound = int(self.processed_df['TOTAL INBOUND CALLS'].sum())
            avg_inbound = float(self.processed_df['TOTAL INBOUND CALLS'].mean())
            
            # Style for summary
            summary_font = Font(bold=True, size=12, color='FFFFFF')
            summary_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # First row: TOTAL INBOUND CALLS
            ws.cell(row=summary_row_1, column=1, value='TOTAL INBOUND CALLS')
            ws.cell(row=summary_row_1, column=1).font = summary_font
            ws.cell(row=summary_row_1, column=1).fill = summary_fill
            ws.cell(row=summary_row_1, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=summary_row_1, column=1).border = border
            
            ws.cell(row=summary_row_1, column=2, value=total_inbound)
            ws.cell(row=summary_row_1, column=2).font = summary_font
            ws.cell(row=summary_row_1, column=2).fill = summary_fill
            ws.cell(row=summary_row_1, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=summary_row_1, column=2).border = border
            
            # Second row: AVERAGE INBOUND CALLS
            ws.cell(row=summary_row_2, column=1, value='AVERAGE INBOUND CALLS')
            ws.cell(row=summary_row_2, column=1).font = summary_font
            ws.cell(row=summary_row_2, column=1).fill = summary_fill
            ws.cell(row=summary_row_2, column=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=summary_row_2, column=1).border = border
            
            ws.cell(row=summary_row_2, column=2, value=f"{avg_inbound:.2f}")
            ws.cell(row=summary_row_2, column=2).font = summary_font
            ws.cell(row=summary_row_2, column=2).fill = summary_fill
            ws.cell(row=summary_row_2, column=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=summary_row_2, column=2).border = border
            
            # Save file
            wb.save(filename)
            
            # Update UI in main thread
            self.root.after(0, lambda: self._excel_export_complete(filename))
            
        except Exception as e:
            error_msg = f"Error creating Excel file: {str(e)}"
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
            self.root.after(0, lambda: self.log(f"ERROR: {error_msg}"))
            
    def _excel_export_complete(self, filename):
        """Called when Excel export is complete"""
        messagebox.showinfo("Success", f"Styled Excel file created: {filename}")
        self.log(f"Styled Excel exported: {filename}")
        self.status_var.set("Excel export completed")
        
    def show_error(self, message):
        """Show error message"""
        messagebox.showerror("Error", message)
        self.status_var.set("Error occurred")
        self.log(f"ERROR: {message}")
        
    def run(self):
        """Run the application"""
        self.log("Agent Performance Data Processor started")
        self.log("Select a CSV file and click 'Process Data' to begin")
        self.root.mainloop()

if __name__ == "__main__":
    app = AgentPerformanceGUI()
    app.run()